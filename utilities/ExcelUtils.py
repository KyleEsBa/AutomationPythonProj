import openpyxl

class ExcelUtils:
    file="C:\\Users\\User\\IdeaProjects\\AutomationPythonProj\\testData\\TestData.xlsx"
    sheetName="Sheet1"

    def openFile(self):
        workbook=openpyxl.load_workbook(self.file)
        sheet=workbook[self.sheetName]
        return sheet

    def getRowCount(self):
        sheet=self.openFile()
        return sheet.max_row
    def getColumnCount(self):
        sheet=self.openFile()
        return sheet.max_column
    def readData(self, rownum, columno):
        sheet=self.openFile()
        return sheet.cell(row=rownum, column=columno).value
    def writeData(self, rownum, columno, data):
        workbook=openpyxl.load_workbook(self.file)
        sheet=workbook[self.sheetName]
        sheet.cell(row=rownum, column=columno).value=data
        workbook.save(self.file)