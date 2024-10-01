import openpyxl

class ExcelUtils:
    file= "../../resources/testData/TestData.xlsx"
    sheetName="Sheet1"

    @staticmethod
    def openFile():
        workbook=openpyxl.load_workbook(ExcelUtils.file)
        sheet=workbook[ExcelUtils.sheetName]
        return sheet

    @staticmethod
    def getRowCount():
        sheet=ExcelUtils.openFile()
        return sheet.max_row

    @staticmethod
    def getColumnCount():
        sheet=ExcelUtils.openFile()
        return sheet.max_column

    @staticmethod
    def readData(rownum, columno):
        sheet=ExcelUtils.openFile()
        return sheet.cell(row=rownum, column=columno).value

    @staticmethod
    def writeData(rownum, columno, data):
        workbook=openpyxl.load_workbook(ExcelUtils.file)
        sheet=workbook[ExcelUtils.sheetName]
        sheet.cell(row=rownum, column=columno).value=data
        workbook.save(ExcelUtils.file)